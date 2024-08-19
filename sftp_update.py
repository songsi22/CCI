import json
import paramiko
from openpyxl import load_workbook
import ssh_command


def ssh_template_update(user, filename, servers):
    ssh = paramiko.SSHClient()
    private_key = paramiko.RSAKey.from_private_key_file('./id_rsa')
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect('serverip', username='user', pkey=private_key)
    sftp = ssh.open_sftp()
    remote_excel_path = f'files/{user}_files/{filename}'
    local_excel_path = f'./{filename}'
    sftp.get(remote_excel_path, local_excel_path)

    wb = load_workbook(local_excel_path)

    systems_info = ssh_command.get_system_info(servers=servers)

    ws = wb.active

    for ip in systems_info.keys():
        search_value = ip
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
            cell = row[0]  # 14번째 열의 셀
            if cell.value == search_value:
                ws[f'B{cell.row}'] = systems_info[ip]['hostname']
                ws[f'C{cell.row}'] = systems_info[ip]['OS']
                ws[f'H{cell.row}'] = systems_info[ip]['swap']
                mount_point_str = "\n".join(systems_info[ip]['MountPoint'])
                ws[f'L{cell.row}'] = mount_point_str
                ws[f'K{cell.row}'] = systems_info[ip]['NASsize']
                IPs_str = "\n".join(systems_info[ip]['IPs'])
                ws[f'N{cell.row}'] = IPs_str
                break  # 값을 찾으면 반복문 종료
    wb.save(local_excel_path)
    wb.close()

    new_remote_excel_path = f'/home/user/CCI/{user}_files/{filename}'
    sftp.put(local_excel_path, new_remote_excel_path)
    sftp.close()
    ssh.close()


with open('remote_comm.json') as f:
    servers = json.load(f)

ssh_template_update(servers[-1]['user'], servers[-2]['filename'], servers[:-2])
'''
remote_comm.json[-1] == user
remote_comm.json[-2] == filename
remote_comm.json[:-2] == server info(hostname,ip,port,user,passwd)
'''