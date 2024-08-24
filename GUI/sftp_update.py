import json
import paramiko
from openpyxl import load_workbook
import ssh_command


def ssh_template_update(user, filename, servers):
    """
    원격 서버에서 템플릿 파일을 다운로드하고, 시스템 정보를 업데이트한 후 다시 업로드하는 함수.

    Args:
        user (str): 원격 서버의 사용자 이름.
        filename (str): 원격 서버에 있는 엑셀 파일의 이름.
        servers (list): 시스템 정보가 포함된 서버 리스트.

    Raises:
        Exception: SSH 연결이나 파일 전송 중 발생한 예외.
    """
    ssh = paramiko.SSHClient()
    private_key = paramiko.RSAKey.from_private_key_file('./id_rsa')  # 키 파일 경로는 환경 변수나 설정 파일로 관리 필요
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        ssh.connect('serverip', username=user, pkey=private_key)  # serverip를 외부에서 주입받도록 개선 필요
        sftp = ssh.open_sftp()
        remote_excel_path = f'/home/{user}/files/{user}_files/{filename}'
        local_excel_path = f'./{filename}'

        # 원격에서 파일 다운로드
        sftp.get(remote_excel_path, local_excel_path)

        wb = load_workbook(local_excel_path)
        ws = wb.active

        systems_info = ssh_command.get_system_info(servers=servers)

        for ip in systems_info.keys():
            search_value = ip
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                cell = row[0]  # 14번째 열의 셀
                if cell.value == search_value:
                    ws[f'B{cell.row}'] = systems_info[ip]['hostname']
                    ws[f'C{cell.row}'] = systems_info[ip]['OS']
                    ws[f'H{cell.row}'] = systems_info[ip]['swap']
                    mount_point_str = "\n".join(systems_info[ip]['MountPoint'])
                    ws[f'L{cell.row}'] = mount_point_str
                    ws[f'K{cell.row}'] = systems_info[ip]['NASsize']
                    IPs_str = "\n".join(systems_info[ip]['IPs'])
                    ws[f'N{cell.row}'] = IPs_str
                    break  # 값을 찾으면 반복문 종료

        # 수정된 파일 저장
        wb.save(local_excel_path)
        wb.close()

        # 수정된 파일 업로드
        new_remote_excel_path = f'/home/{user}/CCI/{user}_files/{filename}'
        sftp.put(local_excel_path, new_remote_excel_path)

    except Exception as e:
        print(f"SSH 연결 또는 파일 전송 중 오류 발생: {e}")
    finally:
        sftp.close()
        ssh.close()


def load_servers_from_file(file_path):
    """
    JSON 파일에서 서버 정보를 로드하는 함수.

    Args:
        file_path (str): 서버 정보가 저장된 JSON 파일의 경로.

    Returns:
        list: 서버 정보가 포함된 리스트.
    """
    with open(file_path) as f:
        return json.load(f)


if __name__ == "__main__":
    servers = load_servers_from_file('remote_comm.json')
    ssh_template_update(servers[-1]['user'], servers[-2]['filename'], servers[:-2])

'''
remote_comm.json[-1] == user
remote_comm.json[-2] == filename
remote_comm.json[:-2] == server info(hostname,ip,port,user,passwd)
'''