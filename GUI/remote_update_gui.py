import os
import sys
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, filedialog
import threading
import json
import paramiko
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import ssh_command


def log_message(label, message):
    """
    상태 라벨에 로그 메시지를 추가하는 함수.
    """
    current_text = label.cget("text")
    updated_text = current_text + message + "\n"
    label.config(text=updated_text)
    label.update()


def resource_path(relative_path):
    """
    PyInstaller와 함께 사용하기 위한 리소스 경로를 반환하는 함수.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def ssh_template_update(ip, port, user, auth_method, key_path, password, base_path, filename, servers, partname, label):
    """
    SSH를 통해 원격 서버에서 템플릿 파일을 업데이트하고 업로드하는 함수.
    """
    cday = datetime.now().strftime("%Y%m%d")
    ctime = datetime.now().strftime("%H%M")
    ssh = paramiko.SSHClient()
    try:
        log_message(label, "연결 설정 중...")

        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        if auth_method == "key":
            private_key = paramiko.RSAKey.from_private_key_file(key_path)
            ssh.connect(ip, port=port, username=user, pkey=private_key)
        else:
            ssh.connect(ip, port=port, username=user, password=password)

        log_message(label, "세션 열기...")
        sftp = ssh.open_sftp()

        customer = filename.split('-')[0]
        csp_type = filename.split('-')[1]
        new_filename = f'{customer}-{csp_type}-inventory-{cday}.xlsx'
        remote_excel_path = f'{base_path}/{partname}_files/{filename}'
        local_excel_path = new_filename
        remote_custom_path = f'{base_path}/{partname}_custom/{customer}'
        local_custom_path = f'./{customer}'

        log_message(label, f"파일 다운로드 중...")
        sftp.get(remote_excel_path, local_excel_path)
        sftp.get(remote_custom_path, local_custom_path)

        log_message(label, "엑셀 파일 로드 중...")
        wb = load_workbook(local_excel_path)
        ws = wb.active

        log_message(label, "시스템 정보 가져오는 중...")
        systems_info = ssh_command.get_system_info(servers=servers)

        log_message(label, "엑셀 파일 업데이트 중...")
        for ip, info in systems_info.items():
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
                cell = row[0]  # 14번째 열의 셀
                if cell.value == ip:
                    ws[f'B{cell.row}'] = info['hostname']
                    ws[f'C{cell.row}'] = info['OS']
                    ws[f'H{cell.row}'] = info['swap']
                    ws[f'L{cell.row}'] = "\n".join(info['MountPoint'])
                    ws[f'K{cell.row}'] = info['NASsize'] if 'NASsize' in info else ''
                    ws[f'N{cell.row}'] = "\n".join(info['IPs'])
                    ws[f'L{cell.row}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    ws[f'K{cell.row}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    ws[f'N{cell.row}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    break

        log_message(label, "엑셀 파일 저장 중...")

        wb.save(local_excel_path)

        with open(f'./{customer}', 'a+', encoding='utf-8') as f:
            f.write(f'{new_filename},{cday}{ctime}\n')

        new_remote_excel_path = f'{base_path}/{partname}_files/{new_filename}'
        log_message(label, f"파일 업로드 중...")
        sftp.put(local_excel_path, new_remote_excel_path)
        sftp.put(local_custom_path, remote_custom_path)

        log_message(label, "작업 완료!")
        messagebox.showinfo("완료", "작업이 완료되었습니다!")

    except Exception as e:
        log_message(label, f"에러 발생: {e}")
        messagebox.showerror("에러", f"작업 중 에러가 발생했습니다: {e}")
    finally:
        log_message(label, "세션 종료 중...")
        sftp.close()
        ssh.close()
        log_message(label, "연결 종료 중...")
        if os.path.exists(local_custom_path):
            os.remove(local_custom_path)


def run_ssh_template_update():
    """
    SSH 템플릿 업데이트 작업을 실행하는 함수.
    """
    ip = ip_entry.get()
    port = int(port_entry.get())
    user = user_entry.get()
    auth_method = auth_method_var.get()
    key_path = key_path_entry.get()
    password = password_entry.get()
    base_path = base_path_entry.get()

    with open('remote_comm.json') as f:
        servers = json.load(f)

    partname = servers[-1]['user']
    filename = servers[-2]['filename']
    server_info = servers[:-2]

    thread = threading.Thread(target=ssh_template_update, args=(
        ip, port, user, auth_method, key_path, password, base_path, filename, server_info, partname, status_label))
    thread.start()


def browse_key_file():
    """
    파일 탐색기에서 키 파일을 선택하는 함수.
    """
    key_file_path = filedialog.askopenfilename(title="Select SSH Private Key File")
    key_path_entry.delete(0, tk.END)
    key_path_entry.insert(0, key_file_path)


# GUI 구성
root = tk.Tk()
root.title("Template Update")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

tk.Label(frame, text="IP 주소:").grid(row=0, column=0, sticky="e")
ip_entry = tk.Entry(frame)
ip_entry.grid(row=0, column=1)

tk.Label(frame, text="포트:").grid(row=1, column=0, sticky="e")
port_entry = tk.Entry(frame)
port_entry.grid(row=1, column=1)
port_entry.insert(0, "22")  # 기본값으로 포트 22 설정

tk.Label(frame, text="사용자 이름:").grid(row=2, column=0, sticky="e")
user_entry = tk.Entry(frame)
user_entry.grid(row=2, column=1)

tk.Label(frame, text="서버 경로:").grid(row=3, column=0, sticky="e")
base_path_entry = tk.Entry(frame)
base_path_entry.grid(row=3, column=1)

auth_method_var = tk.StringVar(value="password")

tk.Label(frame, text="인증 방법:").grid(row=4, column=0, sticky="e")

tk.Radiobutton(frame, text="비밀번호", variable=auth_method_var, value="password").grid(row=4, column=1, sticky="w")
tk.Radiobutton(frame, text="키 파일", variable=auth_method_var, value="key").grid(row=4, column=2, sticky="w")

tk.Label(frame, text="비밀번호:").grid(row=5, column=0, sticky="e")
password_entry = tk.Entry(frame, show="*")
password_entry.grid(row=5, column=1)

tk.Label(frame, text="키 파일 경로:").grid(row=6, column=0, sticky="e")
key_path_entry = tk.Entry(frame)
key_path_entry.grid(row=6, column=1)
browse_button = tk.Button(frame, text="찾기", command=browse_key_file)
browse_button.grid(row=6, column=2)

run_button = tk.Button(frame, text="실행", command=run_ssh_template_update)
run_button.grid(row=7, columnspan=3, pady=10)

status_label = tk.Label(frame, text="", justify="left", anchor="w")
status_label.grid(row=8, columnspan=3, pady=10, sticky="w")

root.mainloop()
