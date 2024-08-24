from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
import string
import os


def data_to_excel(inventories, csp_type, path, cday, customer=''):
    """
    인벤토리 데이터를 엑셀 파일로 저장하는 함수.

    Args:
        inventories (list): 인벤토리 데이터 리스트.
        csp_type (str): CSP 유형.
        path (str): 파일 저장 경로.
        cday (str): 현재 날짜 (YYYYMMDD 형식).
        customer (str, optional): 고객사명. 기본값은 ''.
    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    try:
        wb = load_workbook('files/template.xlsx')
    except FileNotFoundError:
        raise FileNotFoundError("템플릿 파일을 찾을 수 없습니다: files/template.xlsx")

    ws = wb.active
    create_time_in_excel = datetime.now().strftime("%Y년%m월%d일")
    ws['A1'].value = customer
    ws['O1'].value = create_time_in_excel

    for i, inventory in enumerate(inventories):
        ext_ssd = 0
        ext_hdd = 0
        vmguestip = next(iter(inventory))
        inventory = inventory[vmguestip]

        ws[f'A{i + 5}'].value = inventory['availability_zone']  # zone
        ws[f'B{i + 5}'].value = inventory['name']  # name
        ws[f'D{i + 5}'].value = inventory['vcpus']  # vcpu
        ws[f'E{i + 5}'].value = inventory['ram']  # ram

        for volume in inventory['volumes']:
            if volume['bootable']:
                if volume['volume_type'] == 'HDD':
                    ws[f'F{i + 5}'].value = str(volume['size'])  # HDD size
                else:
                    ws[f'G{i + 5}'].value = str(volume['size'])  # SSD size
            else:
                if volume['volume_type'] == 'HDD':
                    ext_hdd += volume['size']
                else:
                    ext_ssd += volume['size']

        if ext_ssd != 0:
            ws[f'J{i + 5}'].value = str(ext_ssd)  # ext SSD total(sum) size
        if ext_hdd != 0:
            ws[f'I{i + 5}'].value = str(ext_hdd)  # ext HDD total(sum) size

        ws[f'M{i + 5}'].value = inventory['publicip']  # pub ip
        ws[f'N{i + 5}'].value = vmguestip  # pri ip
        ws[f'O{i + 5}'].value = inventory['created']  # created
        ws[f'P{i + 5}'].value = inventory['vm_state']  # vm state

        for uppercase in string.ascii_uppercase[:-10]:
            ws[f'{uppercase}{i + 5}'].border = thin_border
            ws[f'{uppercase}{i + 5}'].alignment = Alignment(horizontal='center', vertical='center')

    output_path = f'files/{path}_files/{customer}-{csp_type}-inventory-{cday}.xlsx'

    try:
        wb.save(output_path)
    except Exception as e:
        raise IOError(f"엑셀 파일 저장 중 오류 발생: {str(e)}")


def write_to_file(type, csp_type, customer, path, cday, ctime, filename=''):
    """
        인벤토리 파일 정보를 기록하는 함수.

        Args:
            type (str): 수집 타입 (API 또는 기타).
            csp_type (str): CSP 유형.
            customer (str): 고객사명.
            path (str): 파일 저장 경로.
            cday (str): 현재 날짜 (YYYYMMDD 형식).
            ctime (str): 현재 시간 (HHMM 형식).
            filename (str, optional): 파일 이름. 기본값은 ''.
        """
    customer_path = f'files/{path}_custom/{customer}'
    try:
        with open(customer_path, 'a+', encoding='utf-8') as f:
            if type == 'API':
                record = f'{customer}-{csp_type}-inventory-{cday}.xlsx,{cday}{ctime}\n'
            else:
                record = f'{filename},{cday}{ctime}\n'
            f.write(record)
    except Exception as e:
        raise IOError(f"고객 파일 기록 중 오류 발생: {str(e)}")